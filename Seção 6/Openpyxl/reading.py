# Ler e alterar dados de uma planilha
import openpyxl
from pathlib import Path
from openpyxl import Workbook, load_workbook


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook = load_workbook(WORKBOOK_PATH)

# Nome Para a Planilha
sheet_name = 'My Sheet'

# Selecionou a Planilha
worksheet = workbook[sheet_name]

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value,)


# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)

